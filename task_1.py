# -*- coding: utf-8 -*-
import openpyxl

my_path = ".\\test.xlsx"
wb = openpyxl.load_workbook(my_path)
sheets = wb.sheetnames
ws = wb['Таблица']
ws = wb.active
ws['G1'] = "min"
ws['H1'] = "max"
ws['I1'] = "summ"

# Находим min, max, summ
for i in range(2, 1001):
    val = ws['A{}'.format(i)].value
    val1 = ws['B{}'.format(i)].value
    val2 = ws['C{}'.format(i)].value
    val3 = ws['D{}'.format(i)].value
    val4 = ws['F{}'.format(i)].value
    zeta = [val, val1, val2, val3, val4]
    a = min(zeta)
    b = max(zeta)
    s = sum(zeta)

    ws['G{}'.format(i)] = a
    ws['H{}'.format(i)] = b
    ws['I{}'.format(i)] = s

wb.save('test.xlsx')
