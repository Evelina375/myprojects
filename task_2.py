# -*- coding: utf-8 -*-
import pandas as pd
import sqlite3, csv
import openpyxl
from openpyxl.styles import PatternFill


# считываем значения таблиц

df1 = pd.read_excel('.\\test.xlsx', sheet_name="Справочник", engine='openpyxl')
df2 = pd.read_excel('.\\test.xlsx', sheet_name="Таблица", engine='openpyxl')

# объединяем таблицы
df3 = df1.merge(df2, on='ch', how='right')
df3.to_csv('file1.csv', index = False)
# Создание БД
conn = sqlite3.connect("mydatabase.db")
cursor = conn.cursor()
cursor.execute("DROP TABLE IF EXISTS result")
cursor.execute("""CREATE TABLE result 
                  (ch INT, Name text, One INT,
                   Two INT, Three INT, Four INT, Five INT, min INT, max INT, summ INT)
               """)
# Чтение и запись в БД
with open('file1.csv','r', encoding = 'utf-8-sig') as fin:
    dr = csv.DictReader(fin)
    to_db = [(i['ch'], i['Name'], i['One'], i['Two'], i['Three'],
              i['Four'], i['Five'], i['min'], i['max'], i['summ']) for i in dr]
cursor.executemany("""INSERT INTO result (ch, Name, One,
                   Two, Three, Four, Five, min, max, summ) VALUES
                   (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""", to_db)
conn.commit()
conn.close()

# вывод из БД
con = sqlite3.connect('mydatabase.db')
wb = openpyxl.Workbook()
wb.create_sheet(title='Результат', index=0)
ws = wb.active
ws.append(["ch", "Name", "One", "Two","Three","Four","Five", "min", "max", "summ"])
with con:
    cur = con.cursor()
    cur.execute("SELECT * FROM result")
    rows = cur.fetchall()

[ws.append(row) for row in rows]

con.commit()
con.close()

ws = wb.active
ws.column_dimensions['D'].hidden= True
ws.column_dimensions['E'].hidden= True
ws.column_dimensions['F'].hidden= True
ws.column_dimensions['G'].hidden= True

for row in ws.iter_rows(min_col=3, min_row=2, max_col=3, max_row=1000):
    for cell in row:
        if cell.value >0:
            #print(cell)
            cell.fill = openpyxl.styles.PatternFill(start_color='008000', end_color='008000', fill_type='solid')
            cell.font = openpyxl.styles.Font(color='000000', bold=False, italic=False)
        else:
            cell.fill = openpyxl.styles.PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
            cell.font = openpyxl.styles.Font(color='000000', bold=False, italic=False)

counter = 1
for row in ws.iter_rows(min_col=3, min_row=2, max_col=3, max_row=1000):
    for cell in row:
        d = openpyxl.styles.PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
        if cell.fill.start_color == d.start_color:
            counter += 1
        else:
            counter += 1
            ws.row_dimensions[counter].hidden= True

wb.save('result.xlsx')