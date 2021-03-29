# -*- coding: utf-8 -*-
import openpyxl
import random
from openpyxl.worksheet.table import Table
# создаем новый excel-файл
wb = openpyxl.Workbook()

# добавляем новый лист
wb.create_sheet(title = 'Таблица', index = 0)
wb.create_sheet(title = 'Справочник', index = 1)
ws=wb.active
ws.append(["One", "Two","Three","Four","Five","ch"])
matrix = list()
for i in range(1,1000):
    temp = []
    for j in range(1,6):
        temp.append(random.randint(-100,100))
        matrix.append(temp)
    for j in range(7,8):
        temp.append(random.randint(1,5))
        matrix.append(temp)
    ws.append(temp)
#print(temp)
tab=Table(displayName="Table", ref="A1:F1001")
ws.add_table(tab)

# 2 table
ws=wb['Справочник']
ws.append(["ch", "Name"])
data=[
    [1,'Астрахань'],
    [2,'Волгоград'],
    [3,'Саратов'],
    [4,'Пенза'],
    [5,'Ульяновск'],
]
for row in data:
    ws.append(row)
tab1=Table(displayName="Table2", ref="A1:B5")
ws.add_table(tab1)
wb.save('test.xlsx')